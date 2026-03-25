#!/usr/bin/env python3
"""
코딩 에이전트 릴레이 설치 챌린지 - 문제 생성기
==============================================
5000+ 줄의 복잡한 텍스트 파일과 130개 고유 문제를 생성합니다.
각 문제는 로컬에 설치된 코딩 에이전트만으로 풀 수 있도록 설계되었습니다.
"""

import random
import string
import json
import math
import re
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ============================================================
# 설정
# ============================================================
SEED = 2026
NUM_PARTICIPANTS = 130
PROBLEMS_PER_TYPE = 26  # 5 types x 26 = 130

random.seed(SEED)

# ============================================================
# 이름/단어 풀
# ============================================================
FIRST_NAMES = [
    "James", "Mary", "John", "Patricia", "Robert", "Jennifer", "Michael",
    "Linda", "David", "Elizabeth", "William", "Barbara", "Richard", "Susan",
    "Joseph", "Jessica", "Thomas", "Sarah", "Charles", "Karen", "Daniel",
    "Lisa", "Matthew", "Nancy", "Anthony", "Betty", "Mark", "Margaret",
    "Donald", "Sandra", "Steven", "Ashley", "Paul", "Dorothy", "Andrew",
    "Kimberly", "Joshua", "Emily", "Kenneth", "Donna", "Kevin", "Michelle",
    "Brian", "Carol", "George", "Amanda", "Timothy", "Melissa", "Ronald",
    "Deborah", "Edward", "Stephanie", "Jason", "Rebecca", "Jeffrey", "Sharon",
    "Ryan", "Laura", "Jacob", "Cynthia", "Gary", "Kathleen", "Nicholas",
    "Amy", "Eric", "Angela", "Jonathan", "Shirley", "Stephen", "Anna",
    "Larry", "Brenda", "Justin", "Pamela", "Scott", "Emma", "Brandon",
    "Nicole", "Benjamin", "Helen", "Samuel", "Samantha", "Raymond", "Katherine",
    "Gregory", "Christine", "Frank", "Debra", "Alexander", "Rachel", "Patrick",
    "Carolyn", "Jack", "Janet", "Dennis", "Catherine", "Jerry", "Maria",
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
    "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
    "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark",
    "Ramirez", "Lewis", "Robinson", "Walker", "Young", "Allen", "King",
    "Wright", "Scott", "Torres", "Nguyen", "Hill", "Flores", "Green",
    "Adams", "Nelson", "Baker", "Hall", "Rivera", "Campbell", "Mitchell",
    "Carter", "Roberts", "Gomez", "Phillips", "Evans", "Turner", "Diaz",
    "Parker", "Cruz", "Edwards", "Collins", "Reyes", "Stewart", "Morris",
    "Morales", "Murphy", "Cook", "Rogers", "Gutierrez", "Ortiz", "Morgan",
    "Cooper", "Peterson", "Bailey", "Reed", "Kelly", "Howard", "Ramos",
    "Kim", "Cox", "Ward", "Richardson", "Watson", "Brooks", "Chavez",
    "Wood", "James", "Bennett", "Gray", "Mendoza", "Ruiz", "Hughes",
    "Price", "Alvarez", "Castillo", "Sanders", "Patel", "Myers", "Long",
]

DEPARTMENTS = ["Engineering", "Marketing", "Sales", "HR", "Finance",
               "Operations", "Legal", "Research"]
ROLES = ["Manager", "Senior Engineer", "Junior Engineer", "Analyst",
         "Director", "Intern", "Lead", "Specialist", "Coordinator", "Architect"]
STATUSES = ["active", "inactive", "on_leave", "transferred"]
MODULES = ["auth", "api", "database", "cache", "scheduler", "worker",
           "gateway", "monitor", "storage", "network", "billing", "notifier"]
LOG_LEVELS = ["DEBUG", "INFO", "WARN", "ERROR", "CRITICAL"]
TAGS = ["alpha", "beta", "gamma", "delta", "epsilon", "zeta", "theta", "kappa"]
PROJECTS = ["ProjectAlpha", "ProjectBeta", "ProjectGamma", "ProjectDelta",
            "ProjectEpsilon", "ProjectZeta", "ProjectTheta", "ProjectKappa",
            "ProjectIota", "ProjectLambda"]
CATEGORIES = ["TypeA", "TypeB", "TypeC", "TypeD", "TypeE", "TypeF"]
REGIONS = ["APAC", "EMEA", "AMER", "LATAM", "NORDICS", "MENA"]
PRIORITIES = ["low", "medium", "high", "critical"]
LANGUAGES = ["PYTHON", "JAVASCRIPT", "JAVA", "GO", "RUST"]

FUNC_NAMES_PY = [
    "process_data", "validate_input", "handle_request", "compute_result",
    "transform_output", "parse_config", "check_status", "run_pipeline",
    "fetch_records", "aggregate_metrics", "send_notification", "retry_operation",
    "clean_buffer", "init_session", "close_connection", "verify_token",
    "encode_payload", "decode_response", "filter_results", "merge_datasets",
]
FUNC_NAMES_JS = [
    "processData", "validateInput", "handleRequest", "computeResult",
    "transformOutput", "parseConfig", "checkStatus", "runPipeline",
    "fetchRecords", "aggregateMetrics", "sendNotification", "retryOperation",
    "cleanBuffer", "initSession", "closeConnection", "verifyToken",
    "encodePayload", "decodeResponse", "filterResults", "mergeDatasets",
]
FUNC_NAMES_JAVA = [
    "processData", "validateInput", "handleRequest", "computeResult",
    "transformOutput", "parseConfig", "checkStatus", "runPipeline",
    "fetchRecords", "aggregateMetrics", "sendNotification", "retryOperation",
    "cleanBuffer", "initSession", "closeConnection", "verifyToken",
]
FUNC_NAMES_GO = [
    "ProcessData", "ValidateInput", "HandleRequest", "ComputeResult",
    "TransformOutput", "ParseConfig", "CheckStatus", "RunPipeline",
    "FetchRecords", "AggregateMetrics", "SendNotification", "RetryOperation",
]
FUNC_NAMES_RUST = [
    "process_data", "validate_input", "handle_request", "compute_result",
    "transform_output", "parse_config", "check_status", "run_pipeline",
    "fetch_records", "aggregate_metrics", "send_notification", "retry_operation",
]

CODE_KEYWORDS = ["return", "if", "for", "import", "error", "try", "async", "while",
                 "break", "continue", "match", "yield", "raise", "catch", "throw"]

LOREM_WORDS = [
    "lorem", "ipsum", "dolor", "sit", "amet", "consectetur", "adipiscing",
    "elit", "sed", "do", "eiusmod", "tempor", "incididunt", "ut", "labore",
    "et", "dolore", "magna", "aliqua", "enim", "ad", "minim", "veniam",
    "quis", "nostrud", "exercitation", "ullamco", "laboris", "nisi",
    "aliquip", "ex", "ea", "commodo", "consequat", "duis", "aute", "irure",
    "in", "reprehenderit", "voluptate", "velit", "esse", "cillum", "fugiat",
    "nulla", "pariatur", "excepteur", "sint", "occaecat", "cupidatat",
    "non", "proident", "sunt", "culpa", "qui", "officia", "deserunt",
    "mollit", "anim", "id", "est", "laborum", "system", "server", "data",
    "process", "network", "storage", "memory", "queue", "thread", "module",
    "configuration", "deployment", "infrastructure", "monitoring", "logging",
]


# ============================================================
# 직원 풀 생성
# ============================================================
def generate_employees(n=500):
    employees = []
    for i in range(1, n + 1):
        emp = {
            "employee_id": f"EMP{i:05d}",
            "name": f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}",
            "department": random.choice(DEPARTMENTS),
            "role": random.choice(ROLES),
            "score": round(random.uniform(10, 100), 1),
            "status": random.choices(STATUSES, weights=[60, 15, 15, 10])[0],
            "projects": random.sample(PROJECTS, k=random.randint(1, 4)),
        }
        employees.append(emp)
    return employees


# ============================================================
# 섹션 생성기들
# ============================================================
def generate_log_sections(employees, n_sections=40, lines_per_section=60):
    """로그 섹션 생성. 구조화된 데이터도 함께 반환."""
    sections_text = []
    all_log_entries = []
    base_date = datetime(2025, 1, 1)
    emp_ids = [e["employee_id"] for e in employees]

    for sec_i in range(n_sections):
        section_id = f"LOG-{sec_i + 1:03d}"
        lines = [f"===== LOG_SECTION {section_id} ====="]

        for _ in range(lines_per_section):
            ts = base_date + timedelta(
                days=random.randint(0, 364),
                hours=random.randint(0, 23),
                minutes=random.randint(0, 59),
                seconds=random.randint(0, 59),
            )
            level = random.choices(LOG_LEVELS, weights=[30, 35, 15, 12, 8])[0]
            module = random.choice(MODULES)
            emp_id = random.choice(emp_ids[:300])  # 60% of employees appear in logs
            duration = random.randint(1, 9999)
            tag = random.choice(TAGS)
            msg = " ".join(random.choices(LOREM_WORDS, k=random.randint(3, 8)))

            line = (f"[{ts.strftime('%Y-%m-%dT%H:%M:%SZ')}] [{level}] [{module}] "
                    f"id={emp_id} duration={duration}ms msg=\"{msg}\" tag={tag}")
            lines.append(line)

            all_log_entries.append({
                "timestamp": ts,
                "level": level,
                "module": module,
                "employee_id": emp_id,
                "duration": duration,
                "tag": tag,
                "section_id": section_id,
            })

        lines.append(f"===== END {section_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))

    return sections_text, all_log_entries


def generate_json_blocks(employees, n_blocks=25):
    """JSON 블록 생성."""
    sections_text = []
    all_json_records = []
    random.shuffle(employees)

    chunk_size = len(employees) // n_blocks
    for blk_i in range(n_blocks):
        block_id = f"JSONBLK-{blk_i + 1:03d}"
        lines = [f"===== JSON_BLOCK {block_id} ====="]

        start = blk_i * chunk_size
        end = start + chunk_size if blk_i < n_blocks - 1 else len(employees)
        chunk = employees[start:end]

        for emp in chunk:
            record = {
                "employee_id": emp["employee_id"],
                "name": emp["name"],
                "department": emp["department"],
                "role": emp["role"],
                "score": emp["score"],
                "status": emp["status"],
                "projects": emp["projects"],
                "block_id": block_id,
            }
            # 약간 messy하게 한 줄로 출력 (가끔 여러 줄)
            if random.random() < 0.3:
                lines.append(json.dumps(record, indent=2))
            else:
                lines.append(json.dumps(record))
            all_json_records.append(record)

        lines.append(f"===== END {block_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))

    return sections_text, all_json_records


def generate_csv_tables(n_tables=20, rows_per_table=40):
    """CSV 테이블 생성."""
    sections_text = []
    all_csv_rows = []
    CODE_PREFIXES = ["PRF-A", "PRF-B", "PRF-C", "PRF-D", "PRF-E", "PRF-F"]

    for tbl_i in range(n_tables):
        table_id = f"TBL-{chr(65 + tbl_i // 5)}{tbl_i % 5 + 1:02d}"
        lines = [f"===== CSV_TABLE {table_id} ====="]
        lines.append("table_id,row_num,category,region,value_a,value_b,value_c,flag,code_prefix")

        for row_i in range(rows_per_table):
            cat = random.choice(CATEGORIES)
            region = random.choice(REGIONS)
            va = round(random.uniform(1, 999.99), 2)
            vb = round(random.uniform(1, 999.99), 2)
            vc = round(random.uniform(1, 999.99), 2)
            flag = random.choice([0, 1])
            prefix = random.choice(CODE_PREFIXES)

            # 가끔 빈 셀
            if random.random() < 0.05:
                va = ""
            if random.random() < 0.05:
                vb = ""

            row_data = {
                "table_id": table_id,
                "row_num": row_i + 1,
                "category": cat,
                "region": region,
                "value_a": va,
                "value_b": vb,
                "value_c": vc,
                "flag": flag,
                "code_prefix": prefix,
            }
            line = f"{table_id},{row_i + 1},{cat},{region},{va},{vb},{vc},{flag},{prefix}"
            lines.append(line)
            all_csv_rows.append(row_data)

        lines.append(f"===== END {table_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))

    return sections_text, all_csv_rows


def generate_code_blocks(n_blocks=25):
    """코드 블록 생성."""
    sections_text = []
    all_code_data = []

    lang_func_map = {
        "PYTHON": FUNC_NAMES_PY,
        "JAVASCRIPT": FUNC_NAMES_JS,
        "JAVA": FUNC_NAMES_JAVA,
        "GO": FUNC_NAMES_GO,
        "RUST": FUNC_NAMES_RUST,
    }

    for blk_i in range(n_blocks):
        lang = LANGUAGES[blk_i % len(LANGUAGES)]
        block_id = f"CODE-{blk_i + 1:03d}"
        lines = [f"===== CODE_BLOCK {block_id} [{lang}] ====="]

        funcs_used = []
        code_lines = []
        n_funcs = random.randint(3, 6)
        chosen_funcs = random.sample(lang_func_map[lang], k=min(n_funcs, len(lang_func_map[lang])))

        for func_name in chosen_funcs:
            funcs_used.append(func_name)
            if lang == "PYTHON":
                code_lines.append(f"def {func_name}(data, config=None):")
                for _ in range(random.randint(4, 10)):
                    kw = random.choice(CODE_KEYWORDS + ["pass", "print", "self", "None", "True"])
                    indent = "    "
                    code_lines.append(f"{indent}{kw} {' '.join(random.choices(LOREM_WORDS, k=random.randint(1, 4)))}")
                code_lines.append("")
            elif lang == "JAVASCRIPT":
                code_lines.append(f"function {func_name}(data, options) {{")
                for _ in range(random.randint(4, 10)):
                    kw = random.choice(CODE_KEYWORDS + ["const", "let", "var", "console.log", "null"])
                    code_lines.append(f"  {kw} {' '.join(random.choices(LOREM_WORDS, k=random.randint(1, 4)))};")
                code_lines.append("}")
                code_lines.append("")
            elif lang == "JAVA":
                code_lines.append(f"public void {func_name}(Map<String, Object> data) {{")
                for _ in range(random.randint(4, 10)):
                    kw = random.choice(CODE_KEYWORDS + ["new", "this", "null", "System.out.println", "void"])
                    code_lines.append(f"    {kw} {' '.join(random.choices(LOREM_WORDS, k=random.randint(1, 4)))};")
                code_lines.append("}")
                code_lines.append("")
            elif lang == "GO":
                code_lines.append(f"func {func_name}(data map[string]interface{{}}) error {{")
                for _ in range(random.randint(4, 10)):
                    kw = random.choice(CODE_KEYWORDS + ["fmt.Println", "nil", "err", "range", "defer"])
                    code_lines.append(f"\t{kw} {' '.join(random.choices(LOREM_WORDS, k=random.randint(1, 4)))}")
                code_lines.append("}")
                code_lines.append("")
            elif lang == "RUST":
                code_lines.append(f"fn {func_name}(data: &HashMap<String, Value>) -> Result<(), Error> {{")
                for _ in range(random.randint(4, 10)):
                    kw = random.choice(CODE_KEYWORDS + ["let", "mut", "Ok", "Err", "println!", "match"])
                    code_lines.append(f"    {kw} {' '.join(random.choices(LOREM_WORDS, k=random.randint(1, 4)))}")
                code_lines.append("}")
                code_lines.append("")

        lines.extend(code_lines)
        lines.append(f"===== END {block_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))

        # 각 코드 블록의 메타데이터
        block_content = "\n".join(code_lines)
        all_code_data.append({
            "block_id": block_id,
            "language": lang,
            "functions": funcs_used,
            "content_lines": code_lines,
        })

    return sections_text, all_code_data


def generate_registry_sections(n_sections=12, entries_per_section=30):
    """레지스트리 섹션 생성."""
    sections_text = []
    all_registry_entries = []

    for sec_i in range(n_sections):
        section_id = f"REG-{sec_i + 1:03d}"
        lines = [f"===== REGISTRY {section_id} ====="]

        for ent_i in range(entries_per_section):
            entry_id = f"REG-{sec_i + 1:03d}-{ent_i + 1:04d}"
            tag = random.choice(TAGS)
            priority = random.choice(PRIORITIES)
            ref = f"EMP{random.randint(1, 500):05d}"
            metric = round(random.uniform(0.1, 999.9), 1)

            line = f"{entry_id} | tag={tag} | priority={priority} | ref={ref} | metric={metric}"
            lines.append(line)

            all_registry_entries.append({
                "entry_id": entry_id,
                "section_id": section_id,
                "tag": tag,
                "priority": priority,
                "ref": ref,
                "metric": metric,
            })

        lines.append(f"===== END {section_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))

    return sections_text, all_registry_entries


def generate_prose_blocks(n_blocks=15):
    """산문 블록 (노이즈)."""
    sections_text = []
    for blk_i in range(n_blocks):
        block_id = f"DOC-{blk_i + 1:03d}"
        lines = [f"===== PROSE_BLOCK {block_id} ====="]
        for _ in range(random.randint(10, 25)):
            sentence = " ".join(random.choices(LOREM_WORDS, k=random.randint(8, 20)))
            lines.append(sentence.capitalize() + ".")
        lines.append(f"===== END {block_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))
    return sections_text


def generate_metadata_dumps(n_sections=10):
    """메타데이터 덤프 (노이즈 + 약간의 의미 있는 필드)."""
    sections_text = []
    for sec_i in range(n_sections):
        section_id = f"META-{sec_i + 1:03d}"
        lines = [f"===== METADATA_DUMP {section_id} ====="]
        for _ in range(random.randint(15, 25)):
            # hex 노이즈
            hex_str = "".join(random.choices("0123456789abcdef", k=random.randint(32, 64)))
            if random.random() < 0.3:
                lines.append(f"0x{hex_str} :: checksum={random.randint(1000, 9999)} ref=EMP{random.randint(1, 500):05d}")
            else:
                lines.append(f"0x{hex_str}")
        lines.append(f"===== END {section_id} =====")
        lines.append("")
        sections_text.append("\n".join(lines))
    return sections_text


# ============================================================
# 파일 조립
# ============================================================
def assemble_challenge_file(all_sections):
    """모든 섹션을 섞어서 하나의 텍스트 파일로 조립."""
    random.shuffle(all_sections)

    header = """################################################################
# CHALLENGE DATA FILE - Coding Agent Relay Challenge 2026
# 이 파일은 코딩 에이전트 설치 확인용 챌린지 데이터입니다.
# 총 {}개 섹션이 포함되어 있습니다.
# 주의: 이 파일을 직접 읽어서 문제를 풀려고 하지 마세요.
#       코딩 에이전트에게 이 파일을 분석하도록 요청하세요.
################################################################

""".format(len(all_sections))

    return header + "\n".join(all_sections)


# ============================================================
# 문제 생성기들
# ============================================================
class Problem:
    def __init__(self, pid, ptype, text, answer, params):
        self.pid = pid
        self.ptype = ptype
        self.text = text
        self.answer = answer
        self.params = params


def generate_type_a_problems(log_entries, n=PROBLEMS_PER_TYPE):
    """Type A: 로그 필터링 + 집계"""
    problems = []
    used_answers = set()

    # 파라미터 조합 생성
    combos = []
    for level in LOG_LEVELS:
        for module in MODULES:
            for month in range(1, 13):
                combos.append((level, module, month))
    random.shuffle(combos)

    idx = 0
    for combo in combos:
        if len(problems) >= n:
            break
        level, module, month = combo
        t1 = datetime(2025, month, 1)
        if month == 12:
            t2 = datetime(2025, 12, 31, 23, 59, 59)
        else:
            t2 = datetime(2025, month + 1, 1) - timedelta(seconds=1)

        # 필터링
        filtered = [e for e in log_entries
                    if e["level"] == level
                    and e["module"] == module
                    and t1 <= e["timestamp"] <= t2]

        if len(filtered) < 3:
            continue

        total_duration = sum(e["duration"] for e in filtered)
        count = len(filtered)
        answer = f"{total_duration}"

        if answer in used_answers or total_duration == 0:
            continue

        used_answers.add(answer)
        t1_str = t1.strftime("%Y-%m-%d")
        t2_str = t2.strftime("%Y-%m-%d")

        text = (f"challenge_data.dat 파일에서 다음 조건의 로그 항목을 찾으세요:\n"
                f"- 로그 레벨: {level}\n"
                f"- 모듈: {module}\n"
                f"- 기간: {t1_str} ~ {t2_str}\n"
                f"해당 항목들의 duration 값(ms 제외, 숫자만)을 모두 합산하세요.\n"
                f"(총 {count}개 항목이 있어야 합니다. 개수가 다르면 파싱을 확인하세요)\n"
                f"답: duration 합계 (정수)")

        problems.append(Problem(
            pid=len(problems) + 1,
            ptype="A",
            text=text,
            answer=answer,
            params={"level": level, "module": module, "period": f"{t1_str}~{t2_str}", "count": count}
        ))
        idx += 1

    return problems


def generate_type_b_problems(json_records, log_entries, n=PROBLEMS_PER_TYPE):
    """Type B: JSON-로그 크로스 레퍼런스"""
    problems = []
    used_answers = set()

    combos = []
    for dept in DEPARTMENTS:
        for status in STATUSES:
            for level in ["ERROR", "WARN", "CRITICAL"]:
                combos.append((dept, status, level))
    random.shuffle(combos)

    for combo in combos:
        if len(problems) >= n:
            break
        dept, status, level = combo

        # Step 1: JSON에서 직원 ID 수집
        emp_ids = set()
        for rec in json_records:
            if rec["department"] == dept and rec["status"] == status:
                emp_ids.add(rec["employee_id"])

        if len(emp_ids) < 2:
            continue

        # Step 2: 로그에서 해당 직원 + 레벨 필터링
        filtered_logs = [e for e in log_entries
                         if e["employee_id"] in emp_ids and e["level"] == level]

        if len(filtered_logs) < 2:
            continue

        total_duration = sum(e["duration"] for e in filtered_logs)
        answer = f"{total_duration}"

        if answer in used_answers:
            continue
        used_answers.add(answer)

        text = (f"challenge_data.dat 파일에서 2단계 분석을 수행하세요:\n"
                f"[1단계] JSON_BLOCK 섹션들에서 department가 \"{dept}\"이고 status가 \"{status}\"인 직원들의 employee_id를 수집하세요.\n"
                f"[2단계] LOG_SECTION 섹션들에서 위 employee_id에 해당하면서 로그 레벨이 {level}인 항목들을 찾으세요.\n"
                f"해당 항목들의 duration 값을 모두 합산하세요.\n"
                f"(1단계에서 {len(emp_ids)}명, 2단계에서 {len(filtered_logs)}개 항목이 있어야 합니다)\n"
                f"답: duration 합계 (정수)")

        problems.append(Problem(
            pid=len(problems) + 1,
            ptype="B",
            text=text,
            answer=answer,
            params={"department": dept, "status": status, "level": level,
                    "emp_count": len(emp_ids), "log_count": len(filtered_logs)}
        ))

    return problems


def generate_type_c_problems(csv_rows, n=PROBLEMS_PER_TYPE):
    """Type C: CSV 필터링 + 계산"""
    problems = []
    used_answers = set()

    # 테이블 그룹: TBL-A*, TBL-B*, TBL-C*, TBL-D*
    table_groups = defaultdict(list)
    for row in csv_rows:
        prefix = row["table_id"][:5]  # e.g., "TBL-A"
        table_groups[prefix].append(row)

    combos = []
    for tbl_prefix in sorted(table_groups.keys()):
        for cat in CATEGORIES:
            for region in REGIONS:
                for col in ["value_a", "value_b", "value_c"]:
                    for agg in ["sum", "avg"]:
                        combos.append((tbl_prefix, cat, region, col, agg))
    random.shuffle(combos)

    for combo in combos:
        if len(problems) >= n:
            break
        tbl_prefix, cat, region, col, agg = combo

        filtered = [r for r in table_groups[tbl_prefix]
                    if r["category"] == cat and r["region"] == region
                    and r[col] != ""]

        if len(filtered) < 3:
            continue

        values = [float(r[col]) for r in filtered]
        if agg == "sum":
            result = round(sum(values), 2)
        else:
            result = round(sum(values) / len(values), 2)

        answer = f"{result}"
        if answer in used_answers:
            continue
        used_answers.add(answer)

        agg_kr = "합계" if agg == "sum" else "평균"
        col_display = col.replace("_", " ").upper()

        text = (f"challenge_data.dat 파일에서 CSV 테이블을 분석하세요:\n"
                f"- 대상 테이블: table_id가 \"{tbl_prefix}\"로 시작하는 모든 CSV_TABLE\n"
                f"- 필터 조건: category가 \"{cat}\"이고 region이 \"{region}\"인 행\n"
                f"- 빈 값은 제외하세요\n"
                f"- 계산: 해당 행들의 {col} 열의 {agg_kr}\n"
                f"(총 {len(filtered)}개 행이 매칭되어야 합니다)\n"
                f"답: {agg_kr} 값 (소수점 2자리까지, 예: 123.45)")

        problems.append(Problem(
            pid=len(problems) + 1,
            ptype="C",
            text=text,
            answer=answer,
            params={"table_prefix": tbl_prefix, "category": cat, "region": region,
                    "column": col, "aggregation": agg, "count": len(filtered)}
        ))

    return problems


def generate_type_d_problems(code_data, n=PROBLEMS_PER_TYPE):
    """Type D: 코드 블록 분석"""
    problems = []
    used_answers = set()

    combos = []
    for lang in LANGUAGES:
        for kw in CODE_KEYWORDS:
            combos.append((lang, kw))
    random.shuffle(combos)

    for combo in combos:
        if len(problems) >= n:
            break
        lang, keyword = combo

        # 해당 언어의 모든 코드 블록
        blocks = [b for b in code_data if b["language"] == lang]
        if not blocks:
            continue

        # 키워드를 포함하는 라인 수
        keyword_line_count = 0
        unique_funcs = set()
        for block in blocks:
            for line in block["content_lines"]:
                if keyword in line:
                    keyword_line_count += 1
            unique_funcs.update(block["functions"])

        if keyword_line_count == 0 or len(unique_funcs) == 0:
            continue

        result = keyword_line_count * len(unique_funcs)
        answer = f"{result}"

        if answer in used_answers:
            continue
        used_answers.add(answer)

        text = (f"challenge_data.dat 파일에서 코드 블록을 분석하세요:\n"
                f"- 대상: 언어 태그가 [{lang}]인 모든 CODE_BLOCK\n"
                f"- 분석 1: 해당 블록들의 코드 라인 중 \"{keyword}\"를 포함하는 라인의 수 (A)\n"
                f"- 분석 2: 해당 블록들에서 정의된 고유한 함수/메서드 이름의 수 (B)\n"
                f"  (함수 정의는 def, function, public void, func, fn 으로 시작하는 라인에서 추출)\n"
                f"답: A x B (두 값의 곱, 정수)")

        problems.append(Problem(
            pid=len(problems) + 1,
            ptype="D",
            text=text,
            answer=answer,
            params={"language": lang, "keyword": keyword,
                    "keyword_lines": keyword_line_count, "unique_functions": len(unique_funcs)}
        ))

    return problems


def generate_type_e_problems(registry_entries, n=PROBLEMS_PER_TYPE):
    """Type E: 레지스트리 조회 + 정렬"""
    problems = []
    used_answers = set()

    combos = []
    for tag in TAGS:
        for sort_field in ["metric", "priority"]:
            for return_attr in ["ref", "metric", "entry_id"]:
                combos.append((tag, sort_field, return_attr))
    random.shuffle(combos)

    for combo in combos:
        if len(problems) >= n:
            break
        tag, sort_field, return_attr = combo

        filtered = [e for e in registry_entries if e["tag"] == tag]
        if len(filtered) < 5:
            continue

        if sort_field == "metric":
            filtered.sort(key=lambda x: x["metric"])
        else:
            filtered.sort(key=lambda x: x["priority"])

        # nth item (5~len 사이에서 랜덤)
        max_n = min(len(filtered), 30)
        nth = random.randint(5, max_n)

        item = filtered[nth - 1]  # 1-indexed
        answer = str(item[return_attr])

        if answer in used_answers:
            continue
        used_answers.add(answer)

        sort_kr = "metric (오름차순, 숫자)" if sort_field == "metric" else "priority (오름차순, 알파벳)"
        attr_kr = {"ref": "ref(직원 ID)", "metric": "metric 값", "entry_id": "entry_id"}[return_attr]

        text = (f"challenge_data.dat 파일에서 레지스트리 항목을 분석하세요:\n"
                f"- 대상: 모든 REGISTRY 섹션에서 tag={tag}인 항목을 수집\n"
                f"- 정렬: {sort_kr}\n"
                f"- 결과: 정렬 후 {nth}번째 항목의 {attr_kr}를 제출하세요\n"
                f"(총 {len(filtered)}개 항목이 매칭되어야 합니다)\n"
                f"답: {attr_kr}")

        problems.append(Problem(
            pid=len(problems) + 1,
            ptype="E",
            text=text,
            answer=answer,
            params={"tag": tag, "sort_field": sort_field, "nth": nth,
                    "return_attr": return_attr, "total": len(filtered)}
        ))

    return problems


# ============================================================
# 엑셀 생성
# ============================================================
def create_excel(all_problems, filename="challenge_admin.xlsx"):
    """관리자용 엑셀 생성."""
    wb = Workbook()

    # --- Sheet 1: 마스터 답안지 ---
    ws1 = wb.active
    ws1.title = "마스터 답안지"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["참가자번호", "조번호", "문제유형", "문제", "정답", "제출답안", "결과", "검증용 파라미터"]
    for col_i, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for i, p in enumerate(all_problems):
        row = i + 2
        group_num = (i // 7) + 1
        ws1.cell(row=row, column=1, value=i + 1).border = thin_border
        ws1.cell(row=row, column=2, value=group_num).border = thin_border
        ws1.cell(row=row, column=3, value=p.ptype).border = thin_border
        cell_q = ws1.cell(row=row, column=4, value=p.text)
        cell_q.alignment = Alignment(wrap_text=True)
        cell_q.border = thin_border
        ws1.cell(row=row, column=5, value=p.answer).border = thin_border

        # 제출답안 (입력용, 노란 배경)
        cell_input = ws1.cell(row=row, column=6, value="")
        cell_input.fill = input_fill
        cell_input.border = thin_border

        # 결과 수식
        cell_result = ws1.cell(row=row, column=7)
        cell_result.value = f'=IF(F{row}="","대기",IF(TRIM(TEXT(F{row},"@"))=TRIM(TEXT(E{row},"@")),"PASS","FAIL"))'
        cell_result.border = thin_border

        ws1.cell(row=row, column=8, value=json.dumps(p.params, ensure_ascii=False)).border = thin_border

    # 조건부 서식
    ws1.conditional_formatting.add(
        f"G2:G{len(all_problems) + 1}",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=pass_fill)
    )
    ws1.conditional_formatting.add(
        f"G2:G{len(all_problems) + 1}",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=fail_fill)
    )

    # 열 너비
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 80
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 50

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:H{len(all_problems) + 1}"

    # --- Sheet 2: 문제카드 배포용 ---
    ws2 = wb.create_sheet("문제카드 배포용")

    card_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    card_headers = ["참가자번호", "조번호", "챌린지 문제"]
    for col_i, header in enumerate(card_headers, 1):
        cell = ws2.cell(row=1, column=col_i, value=header)
        cell.font = header_font
        cell.fill = card_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, p in enumerate(all_problems):
        row = i + 2
        group_num = (i // 7) + 1
        ws2.cell(row=row, column=1, value=i + 1).border = thin_border
        ws2.cell(row=row, column=2, value=group_num).border = thin_border
        cell_q = ws2.cell(row=row, column=3, value=p.text)
        cell_q.alignment = Alignment(wrap_text=True)
        cell_q.border = thin_border

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 100
    ws2.freeze_panes = "A2"

    # --- Sheet 3: 진행 가이드 ---
    ws3 = wb.create_sheet("진행 가이드")
    guide_content = [
        ["코딩 에이전트 릴레이 설치 챌린지 - 진행 가이드", ""],
        ["", ""],
        ["[개요]", ""],
        ["목적", "팀원 전원이 코딩 에이전트(Roo Code 또는 Open Code)를 설치하고 사용할 수 있는지 확인"],
        ["인원", "약 130명 (6~7명 x 약 19조)"],
        ["방식", "릴레이 - 첫 번째 사람이 설치 후 다음 사람에게 설치 방법을 전수"],
        ["", ""],
        ["[사전 준비]", ""],
        ["1", "generate_challenge.py를 실행하여 challenge_data.dat와 이 엑셀 파일 생성"],
        ["2", "challenge_data.dat를 참가자 전원에게 배포 (공유 드라이브, Slack 등)"],
        ["3", "각 참가자에게 참가자번호(1~130) 할당"],
        ["4", "'문제카드 배포용' 시트에서 각자의 문제를 확인하도록 안내"],
        ["", ""],
        ["[Roo Code 설치 방법]", ""],
        ["1", "VS Code에서 Extensions(Ctrl+Shift+X) 열기"],
        ["2", "'Roo Code' 검색 후 Install 클릭"],
        ["3", "사이드바에서 Roo Code 아이콘 클릭하여 활성화"],
        ["4", "API 키 설정 (필요한 경우)"],
        ["", ""],
        ["[Open Code 설치 방법]", ""],
        ["1", "터미널에서: npm install -g @anthropic-ai/opencode"],
        ["2", "또는: npx @anthropic-ai/opencode"],
        ["3", "API 키 설정"],
        ["", ""],
        ["[챌린지 진행]", ""],
        ["1", "코딩 에이전트를 설치합니다"],
        ["2", "challenge_data.dat를 작업 폴더에 저장합니다"],
        ["3", "코딩 에이전트에게 자신의 문제를 그대로 입력합니다"],
        ["4", "에이전트가 코드를 작성하고 실행하여 답을 알려줍니다"],
        ["5", "답을 운영자에게 제출합니다"],
        ["", ""],
        ["[문제 유형 설명]", ""],
        ["Type A", "로그 필터링 + 집계: 특정 조건의 로그 항목을 찾아 duration 합산"],
        ["Type B", "크로스 레퍼런스: JSON에서 직원 ID 수집 → 로그에서 해당 ID 필터링 → 집계"],
        ["Type C", "CSV 필터링 + 계산: 특정 조건의 CSV 행을 찾아 수치 계산"],
        ["Type D", "코드 블록 분석: 특정 언어 블록에서 키워드/함수 수 분석"],
        ["Type E", "레지스트리 조회: 태그 필터 → 정렬 → N번째 항목 추출"],
        ["", ""],
        ["[답 검증]", ""],
        ["", "'마스터 답안지' 시트의 F열에 제출 답안을 입력하면 G열에서 자동 판정됩니다"],
        ["", "PASS = 정답, FAIL = 오답, 대기 = 미제출"],
    ]

    title_font = Font(bold=True, size=14, color="4472C4")
    section_font = Font(bold=True, size=11, color="70AD47")

    for row_i, (col_a, col_b) in enumerate(guide_content, 1):
        cell_a = ws3.cell(row=row_i, column=1, value=col_a)
        cell_b = ws3.cell(row=row_i, column=2, value=col_b)
        if row_i == 1:
            cell_a.font = title_font
        elif col_a.startswith("[") and col_a.endswith("]"):
            cell_a.font = section_font

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 80

    # --- Sheet 4: 진행현황 대시보드 ---
    ws4 = wb.create_sheet("진행현황")
    dash_headers = ["조번호", "총인원", "완료", "미완료", "완료율"]
    for col_i, header in enumerate(dash_headers, 1):
        cell = ws4.cell(row=1, column=col_i, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    total_groups = math.ceil(NUM_PARTICIPANTS / 7)
    for g in range(1, total_groups + 1):
        row = g + 1
        start_row = (g - 1) * 7 + 2
        end_row = min(g * 7 + 1, NUM_PARTICIPANTS + 1)

        ws4.cell(row=row, column=1, value=g).border = thin_border
        ws4.cell(row=row, column=2, value=min(7, NUM_PARTICIPANTS - (g - 1) * 7)).border = thin_border
        ws4.cell(row=row, column=3,
                 value=f'=COUNTIF(마스터 답안지!G{start_row}:G{end_row},"PASS")').border = thin_border
        ws4.cell(row=row, column=4,
                 value=f'=B{row}-C{row}').border = thin_border
        ws4.cell(row=row, column=5,
                 value=f'=IF(B{row}=0,0,C{row}/B{row})').border = thin_border
        ws4.cell(row=row, column=5).number_format = '0%'

    # 전체 합계
    total_row = total_groups + 2
    ws4.cell(row=total_row, column=1, value="전체").font = Font(bold=True)
    ws4.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_groups + 1})")
    ws4.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_groups + 1})")
    ws4.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_groups + 1})")
    ws4.cell(row=total_row, column=5, value=f"=IF(B{total_row}=0,0,C{total_row}/B{total_row})")
    ws4.cell(row=total_row, column=5).number_format = '0%'

    for col_i in range(1, 6):
        ws4.cell(row=total_row, column=col_i).font = Font(bold=True)
        ws4.cell(row=total_row, column=col_i).border = thin_border

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 12
    ws4.freeze_panes = "A2"

    wb.save(filename)
    print(f"  -> {filename} 저장 완료")


# ============================================================
# 메인
# ============================================================
def main():
    print("=" * 60)
    print("코딩 에이전트 릴레이 설치 챌린지 - 문제 생성기")
    print("=" * 60)

    # 1. 데이터 생성
    print("\n[1/6] 직원 풀 생성...")
    employees = generate_employees(500)

    print("[2/6] 섹션 생성...")
    log_sections, log_entries = generate_log_sections(employees)
    json_sections, json_records = generate_json_blocks(employees)
    csv_sections, csv_rows = generate_csv_tables()
    code_sections, code_data = generate_code_blocks()
    registry_sections, registry_entries = generate_registry_sections()
    prose_sections = generate_prose_blocks()
    meta_sections = generate_metadata_dumps()

    all_sections = (log_sections + json_sections + csv_sections +
                    code_sections + registry_sections + prose_sections + meta_sections)

    print(f"  총 {len(all_sections)}개 섹션 생성")
    print(f"  - 로그: {len(log_sections)} ({len(log_entries)} entries)")
    print(f"  - JSON: {len(json_sections)} ({len(json_records)} records)")
    print(f"  - CSV: {len(csv_sections)} ({len(csv_rows)} rows)")
    print(f"  - 코드: {len(code_sections)} ({len(code_data)} blocks)")
    print(f"  - 레지스트리: {len(registry_sections)} ({len(registry_entries)} entries)")
    print(f"  - 산문: {len(prose_sections)} blocks")
    print(f"  - 메타데이터: {len(meta_sections)} dumps")

    # 2. 파일 조립
    print("\n[3/6] challenge_data.dat 생성...")
    file_content = assemble_challenge_file(all_sections)
    with open("challenge_data.dat", "w", encoding="utf-8") as f:
        f.write(file_content)
    line_count = file_content.count("\n") + 1
    file_size = len(file_content.encode("utf-8"))
    print(f"  -> {line_count:,} 줄, {file_size:,} bytes ({file_size / 1024:.1f} KB)")

    # 3. 문제 생성
    print("\n[4/6] 130개 고유 문제 생성...")
    problems_a = generate_type_a_problems(log_entries)
    problems_b = generate_type_b_problems(json_records, log_entries)
    problems_c = generate_type_c_problems(csv_rows)
    problems_d = generate_type_d_problems(code_data)
    problems_e = generate_type_e_problems(registry_entries)

    print(f"  Type A (로그 집계): {len(problems_a)}개")
    print(f"  Type B (크로스레퍼런스): {len(problems_b)}개")
    print(f"  Type C (CSV 계산): {len(problems_c)}개")
    print(f"  Type D (코드 분석): {len(problems_d)}개")
    print(f"  Type E (레지스트리 조회): {len(problems_e)}개")

    # 문제 배정: 타입을 순환하며 배정 (조 내에서 다양한 유형이 나오도록)
    all_pools = [problems_a, problems_b, problems_c, problems_d, problems_e]
    pool_indices = [0, 0, 0, 0, 0]
    all_problems = []

    for i in range(NUM_PARTICIPANTS):
        type_idx = i % 5
        pool = all_pools[type_idx]
        if pool_indices[type_idx] < len(pool):
            p = pool[pool_indices[type_idx]]
            p.pid = i + 1
            all_problems.append(p)
            pool_indices[type_idx] += 1
        else:
            # 풀이 부족하면 다른 타입에서 가져옴
            for alt in range(5):
                if pool_indices[alt] < len(all_pools[alt]):
                    p = all_pools[alt][pool_indices[alt]]
                    p.pid = i + 1
                    all_problems.append(p)
                    pool_indices[alt] += 1
                    break

    # 답 중복 검사
    answers = [p.answer for p in all_problems]
    unique_answers = set(answers)
    print(f"\n  총 문제: {len(all_problems)}개")
    print(f"  고유 답: {len(unique_answers)}개")
    if len(unique_answers) < len(all_problems):
        dup_count = len(all_problems) - len(unique_answers)
        print(f"  ⚠ 중복 답: {dup_count}개 (타입 간 중복은 허용)")

    # 4. 엑셀 생성
    print("\n[5/6] 엑셀 파일 생성...")
    create_excel(all_problems, "challenge_admin.xlsx")

    # 5. 완료 요약
    print("\n[6/6] 완료!")
    print("=" * 60)
    print("생성된 파일:")
    print(f"  1. challenge_data.dat  ({line_count:,} 줄, {file_size / 1024:.1f} KB)")
    print(f"  2. challenge_admin.xlsx (4개 시트)")
    print("=" * 60)
    print("\n사용 방법:")
    print("  1. challenge_data.dat를 참가자에게 배포")
    print("  2. challenge_admin.xlsx의 '문제카드 배포용' 시트에서 각자 문제 확인")
    print("  3. 참가자가 코딩 에이전트로 문제를 풀고 답 제출")
    print("  4. '마스터 답안지' 시트의 F열에 제출답안 입력 → G열에서 자동 판정")


def get_all_problems():
    """외부에서 호출하여 130개 문제 목록을 반환하는 함수."""
    random.seed(SEED)
    employees = generate_employees(500)
    log_sections, log_entries = generate_log_sections(employees)
    json_sections, json_records = generate_json_blocks(employees)
    csv_sections, csv_rows = generate_csv_tables()
    code_sections, code_data = generate_code_blocks()
    registry_sections, registry_entries = generate_registry_sections()
    _ = generate_prose_blocks()
    _ = generate_metadata_dumps()

    problems_a = generate_type_a_problems(log_entries)
    problems_b = generate_type_b_problems(json_records, log_entries)
    problems_c = generate_type_c_problems(csv_rows)
    problems_d = generate_type_d_problems(code_data)
    problems_e = generate_type_e_problems(registry_entries)

    all_pools = [problems_a, problems_b, problems_c, problems_d, problems_e]
    pool_indices = [0, 0, 0, 0, 0]
    all_problems = []

    for i in range(NUM_PARTICIPANTS):
        type_idx = i % 5
        pool = all_pools[type_idx]
        if pool_indices[type_idx] < len(pool):
            p = pool[pool_indices[type_idx]]
            p.pid = i + 1
            all_problems.append(p)
            pool_indices[type_idx] += 1
        else:
            for alt in range(5):
                if pool_indices[alt] < len(all_pools[alt]):
                    p = all_pools[alt][pool_indices[alt]]
                    p.pid = i + 1
                    all_problems.append(p)
                    pool_indices[alt] += 1
                    break

    return all_problems


if __name__ == "__main__":
    main()
